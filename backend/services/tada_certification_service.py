"""
TADA 認證文件查詢的多輪追問機制（依據《TADA客服機器人_認證問題集_20260810.xlsx》規則設計）。

範圍說明：這裡只實作「偵測文件類型關鍵字 → 判斷缺市場/缺遊戲 → 用 ForceReply 追問 →
比對回覆 → 逾時換句話重問 → 二次逾時轉人工」這一整套多輪對話機制，並用文件裡「附錄：
文件類型關鍵字索引」的完整資料做為可測試的真實情境。

語言：追問/回覆的語言在第一次偵測到關鍵字的當下就依原始問句判斷一次（中文字元 vs. 純英文），
並存進 pending 狀態的 is_chinese 欄位，後續整段追問（包含逾時重問）都沿用同一個語言，
不會因為使用者回覆的內容（例如純市場英文名稱、純數字 GameID）而中途變成另一種語言。

資料來源：各市場的 Game List Google 試算表網址由使用者提供，實際分頁名稱/gid/欄位配置
已逐一開啟核對過，跟文件《附錄A》描述的欄位配置一致。Brazil 是唯一例外，資料不在
「Brazils Certification List」分頁（該分頁另有其他用途），而是使用者另外指出的分頁
（gid=1216924874），結構也跟其他市場不同：前幾列是不分遊戲的 RGS/RNG 平台認證，
其餘列才是逐遊戲的認證資料。

`build_final_reply()` 現在會直接讀取對應市場的 Game List 試算表查表（24 小時快取，
匯出網址跟 tada_gamelist_service.py 一樣是公開匯出、不需 API Key，差別是這裡改用
export?format=xlsx（而非 csv）＋ openpyxl 讀取，因為認證欄位存的是「檔名文字 + 儲存格
超連結」，CSV 匯出只會保留檔名文字、超連結會遺失，XLSX 用 openpyxl 讀 cell.hyperlink.target
才能把真正的 Google Drive 連結一併帶出來），找不到遊戲/該欄位沒有資料時，如實回覆查無
資料並轉人工，不會用猜的。
"""
import asyncio
import io
import json
import logging
import re
import time
from datetime import datetime, timedelta

import openpyxl
import requests

logger = logging.getLogger(__name__)

PENDING_TTL_MINUTES = 30
_TABLE_CACHE_TTL_SECONDS = 24 * 3600

_ZH_RE = re.compile(r'[一-鿿㐀-䶿]')

# 「certificate file」涵蓋的市場清單：也是 certificate/certification 等泛稱詞的預設對應
# （見下方 _GENERIC_CERT_WORDS 說明），獨立成常數避免兩處清單各自維護、日後改市場清單漏改一邊。
_CERTIFICATE_FILE_MARKETS = ["Greece", "Italy", "Malta", "Netherlands", "Portugal", "Romania", "South Africa",
                             "Spain", "Sweden", "UK", "Brazil", "Belarus", "Ukraine"]

# 文件類型關鍵字 → (對應市場清單, 文件性質)
# 文件性質："game"=需指定遊戲才能查、"platform"=平台層級不分遊戲、"vendor_escalate"=依廠商各自產生，不追問直接轉人工
_DOC_KEYWORD_RULES = (
    (("certificate file", "certification authority"), _CERTIFICATE_FILE_MARKETS, "game"),
    (("hgc approval letter",), ["Greece"], "game"),
    (("symbol mapping",), ["Portugal"], "game"),
    (("help files",), ["Romania", "South Africa", "Spain"], "game"),
    (("dgoj approval", "homologation report", "homologation"), ["Spain"], "game"),
    (("ukgc registration",), ["UK"], "game"),
    (("rgs certificate", "rng certificate"), ["Brazil"], "platform"),
    (("branded report",), ["Brazil", "Italy"], "vendor_escalate"),
    (("game project", "game rule+screeshots"), ["Italy"], "game"),
    (("resolucion directorial", "mincetur approval letter"), ["Peru"], "game"),
)

# 5d：關鍵字命中後市場已固定，但實際核對過資料來源整份表格皆為空，直接轉人工、不追問
# 「malta certificate」原本也在這份清單裡，但核對後發現 Malta 市場實際上有對應的 Game List
# 跟認證連結（跟真的查無資料的 malta license 不同），已移除，改走正常的市場/文件比對流程。
_NO_DATA_KEYWORDS = (
    "malta license", "gli 19", "colombia certificate", "ukgc certificate",
)

# certificate/certification（含中文「認證」）本身就代表在問認證相關的事，但沒有具體到能對到
# 上面任何一組關鍵字時，實際情境幾乎都是在問最常見、涵蓋最多市場的「certificate file」（逐遊戲
# 認證文件），而不是真的完全無法判斷——所以 detect_doc_query 直接把這幾個泛稱詞當成
# certificate file 處理（見下方），只需追問「市場」即可，不會像先前那樣要求使用者在單次回覆裡
# 重新講出具體詞組才能往下走，避免「Can you provide the game certificate」這種很口語的問法
# 一路掉到查無資料轉人工。這幾個字本身已經包含在 _DOC_KEYWORD_RULES 的具體詞組裡（例如
# "certificate file"），所以 detect_doc_query 會先比對具體詞組，比對不到才會落到這裡的泛稱判斷。
_GENERIC_CERT_WORDS = ("certification", "certificate", "認證", "认证")

_MARKET_ALIASES = {
    "Greece": ["greece", "希臘", "希腊"],
    "Italy": ["italy", "義大利", "意大利"],
    "Malta": ["malta", "馬爾他", "马耳他"],
    "Netherlands": ["netherlands", "荷蘭", "荷兰"],
    "Portugal": ["portugal", "葡萄牙"],
    "Romania": ["romania", "羅馬尼亞", "罗马尼亚"],
    "South Africa": ["south africa", "南非"],
    "Spain": ["spain", "西班牙"],
    "Sweden": ["sweden", "瑞典"],
    "UK": ["uk", "united kingdom", "英國", "英国"],
    "Brazil": ["brazil", "巴西"],
    "Peru": ["peru", "秘魯", "秘鲁"],
    "Belarus": ["belarus", "白俄羅斯", "白俄罗斯"],
    "Ukraine": ["ukraine", "烏克蘭", "乌克兰"],
}

# Belarus/Ukraine 沒有自己專屬的認證欄位，是跟其他 11 個國家共用同一組欄位（Y欄起）存放，
# 目前解析不出「廠商問的是哪一份文件」，與其亂猜，不如直接把整份 Game List 連結給廠商自己查。
_SHARED_REPO_MARKETS = {
    "Belarus": "https://docs.google.com/spreadsheets/d/1tP3ax65heTM_njAHz-INDWEIEYFIZ-4pb7LUN0VlRaU/edit",
    "Ukraine": "https://docs.google.com/spreadsheets/d/1S-NQCvf8Re3_4Wcf53CphMv3sxxSHxDiVRImioQ6ck4/edit",
}


def is_shared_repo_market(market: str) -> bool:
    return market in _SHARED_REPO_MARKETS


def shared_repo_reply(market: str, is_chinese: bool) -> str:
    link = _SHARED_REPO_MARKETS.get(market, "")
    if is_chinese:
        return f"{market} 的認證資料是跟其他市場共用同一份資料表，無法單獨查詢特定文件，請直接參考完整 Game List：\n{link}"
    return (f"{market}'s certification data is shared with other markets in one combined sheet and can't be "
            f"looked up individually, please check the full Game List directly:\n{link}")

# 各市場 Game List 試算表：sheet_id/gid 已逐一開啟核對過分頁名稱與欄位配置。
# folder_link 對應文件《附錄A》的「備用資料夾連結」，沒有的市場填 None，
# 廠商要求「整個市場」資料時改給 gamelist_link（該市場整份 Game List 連結）。
_MARKET_DATA_SOURCES = {
    "Greece": {
        "sheet_id": "1GBX3rh9M6DCAoQkdNrF-eshvZUnkyXw4aMie4ocwwm0", "gid": "628814297",
        "folder_link": "https://drive.google.com/drive/folders/1AR44ecaIuhUKeNUT39wjO0vxWkD39OX0",
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1GBX3rh9M6DCAoQkdNrF-eshvZUnkyXw4aMie4ocwwm0/edit",
    },
    "Italy": {
        "sheet_id": "1ASA1I9XXWMRtzcaj9Kzp-OS8p6BCgszexKa94Upj8sQ", "gid": "131560561",
        "folder_link": None,
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1ASA1I9XXWMRtzcaj9Kzp-OS8p6BCgszexKa94Upj8sQ/edit",
    },
    "Malta": {
        "sheet_id": "1Sx5iinHhCej46Q1QzkN9RIIzFK9cnlCub-BhSpScw5A", "gid": "1839087692",
        "folder_link": None,
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1Sx5iinHhCej46Q1QzkN9RIIzFK9cnlCub-BhSpScw5A/edit",
    },
    "Netherlands": {
        "sheet_id": "1cQ-orv4yekG86BxkJ8bV3y_1I-E2ism68UEXMkEPYS8", "gid": "0",
        "folder_link": None,
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1cQ-orv4yekG86BxkJ8bV3y_1I-E2ism68UEXMkEPYS8/edit",
    },
    "Portugal": {
        "sheet_id": "1gYTzt-SdxYaQQPLmQqNHCsZxolBark1-r54GftuNluc", "gid": "1839087692",
        "folder_link": "https://drive.google.com/drive/folders/1zUiJhX5EURm59k5X_BW1kHWbvYFpuciH",
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1gYTzt-SdxYaQQPLmQqNHCsZxolBark1-r54GftuNluc/edit",
    },
    "Romania": {
        "sheet_id": "1-2WlyJ-jx-u8_7GYNQpkLMXHsJYiWTsJF0gju-A58hc", "gid": "1839087692",
        "folder_link": "https://drive.google.com/drive/folders/1gHo0-0JOtmMpcuLZqYQZQ3NZN2kTPbz5",
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1-2WlyJ-jx-u8_7GYNQpkLMXHsJYiWTsJF0gju-A58hc/edit",
    },
    "South Africa": {
        "sheet_id": "1jQXPhOMbXWtCUn7mSXhJvk3lGR69y6cZfBIZVixPVMc", "gid": "628814297",
        "folder_link": "https://drive.google.com/drive/folders/1WMWxYxVudXEI5YXlBFzP3t55BDebz6cq",
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1jQXPhOMbXWtCUn7mSXhJvk3lGR69y6cZfBIZVixPVMc/edit",
    },
    "Spain": {
        "sheet_id": "1Z9IAC2EMo3lWW83itWLxs57ujt1NjlWQkVZLkZ1kRUM", "gid": "1839087692",
        "folder_link": "https://drive.google.com/drive/folders/1AzowiyfAJ1GjSYwI6Fz1eUGsEy5Bi5ow",
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1Z9IAC2EMo3lWW83itWLxs57ujt1NjlWQkVZLkZ1kRUM/edit",
    },
    "Sweden": {
        "sheet_id": "1GNYw0L4bJVFjGs9qHHyOWUt4kk-s2nw2F9EU9ynKTBY", "gid": "677919004",
        "folder_link": None,
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1GNYw0L4bJVFjGs9qHHyOWUt4kk-s2nw2F9EU9ynKTBY/edit",
    },
    "UK": {
        "sheet_id": "1GndNGTYY1igep34458SMZ5_1Ox_65e99MHgr4Wq-kCs", "gid": "1839087692",
        "folder_link": None,
        "gamelist_link": "https://docs.google.com/spreadsheets/d/1GndNGTYY1igep34458SMZ5_1Ox_65e99MHgr4Wq-kCs/edit",
    },
}

_BRAZIL_SHEET_ID = "1XgsIpYWwAZTHfdX7cS-_SPoas876s4YIl4brMMD8ZrE"
_BRAZIL_GID = "1216924874"

# 命中的文件關鍵字 → 要在 Game List 分頁裡找哪一欄（比對 normalize 過的表頭子字串）
_KEYWORD_TO_COLUMN_HINT = {
    "certificate file": "certificate file",
    "certification authority": "certificate file",  # 通用詞：直接給 certificate file（+一併附上 authority）
    "hgc approval letter": "hgc approval letter",
    "symbol mapping": "symbol mapping",
    "help files": "help files",
    "dgoj approval": "dgoj",
    "homologation report": "dgoj",
    "homologation": "dgoj",
    "ukgc registration": "ukgc registration",
}

# 「要整個市場資料，不是單一遊戲」的措辭 → 命中時直接給市場層級連結，不追問遊戲（見文件 5a）
_WHOLE_MARKET_PHRASES = (
    "whole market", "entire market", "full list", "total list", "not just one game", "not just a single game",
    "整個市場", "整个市场", "全部遊戲", "全部游戏", "總表", "总表",
)

# (市場, 文件關鍵字) → 這個文件性質「針對遊戲」但目前完全沒有逐遊戲資料，只能給整個市場的
# 資料夾連結（文件 5a 規則②：「沒有逐遊戲資料時，誠實告知並退回整個市場的資料夾連結」）。
# 這裡的連結是各自獨立的專屬資料夾，跟該市場主要認證的備用資料夾連結不是同一個。
_NO_PER_GAME_DATA = {
    ("Spain", "help files"): "https://drive.google.com/drive/folders/1wu00tHoXEf0TmqRnG13QBHswN2qXbtwX",
    ("Italy", "game project"): "https://drive.google.com/drive/folders/1BDXoe8CsKvgMyNCqLqvDmCEeXK5pZveV",
    ("Italy", "game rule+screeshots"): "https://drive.google.com/drive/folders/1BDXoe8CsKvgMyNCqLqvDmCEeXK5pZveV",
    ("Peru", "resolucion directorial"): "https://drive.google.com/drive/folders/14y6kLreVW_oljFNCqh6ydolF67qEhPYP",
    ("Peru", "mincetur approval letter"): "https://drive.google.com/drive/folders/14y6kLreVW_oljFNCqh6ydolF67qEhPYP",
}

_GAME_ID_INLINE_RE = re.compile(r'game\s*id\s*[:#]?\s*(\d+)', re.IGNORECASE)


def extract_inline_game_id(text: str):
    """從原始問句裡直接找有沒有已經帶 GameID（例如「for GameID 659」），
    有的話就不用再追問遊戲——廠商一次講清楚的問題不該被迫多問一輪。
    只認「GameID + 數字」這種明確格式，不猜測裸數字或遊戲名稱，避免誤判。"""
    m = _GAME_ID_INLINE_RE.search(text)
    return m.group(1) if m else None

_market_table_cache = {}
_brazil_table_cache = {"fetched_at": 0.0}


def _normalize_header(h: str) -> str:
    return re.sub(r'\s+', ' ', h or "").strip().lower()


def _cell_text(value) -> str:
    """openpyxl 讀出的數值型儲存格（例如 GameID）是 float，轉成字串時去掉多餘的 .0。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _extract_row(ws_row) -> list:
    """把一列 openpyxl Cell 轉成 (文字, 超連結網址或None) 的 tuple 清單。"""
    return [(_cell_text(c.value), c.hyperlink.target if c.hyperlink else None) for c in ws_row]


def _find_col(headers: list, substr: str):
    for i, h in enumerate(headers):
        if substr in h:
            return i
    return None


def _find_row(rows: list, id_col, name_col, identifier: str):
    """先用 GameID（純數字比對）找，找不到再用遊戲名稱（先完全相符，再退而找子字串）。
    子字串比對時如果同時符合超過一款遊戲（例如問「Charge Buffalo」，清單裡剛好同時有
    「3 Charge Buffalo」跟「Charge Buffalo-ASCENT」），寧可當作找不到、交由呼叫端誠實
    回覆查無資料，也不要隨便挑第一個猜——猜錯會直接給廠商錯誤的證書，比「答不出來」更糟。"""
    ident = identifier.strip()
    digits = re.sub(r'\D', '', ident)
    if id_col is not None and digits:
        for row in rows:
            if id_col < len(row) and row[id_col][0] == digits:
                return row
    if name_col is not None:
        lower_ident = ident.lower()
        for row in rows:
            if name_col < len(row) and row[name_col][0].lower() == lower_ident:
                return row
        if lower_ident:
            substring_matches = [
                row for row in rows
                if name_col < len(row) and lower_ident in row[name_col][0].lower()
            ]
            if len(substring_matches) == 1:
                return substring_matches[0]
    return None


def _fetch_xlsx(sheet_id: str, gid: str):
    """公開匯出網址不需登入/不需 API Key，改用 xlsx 格式是為了保留儲存格超連結
    （csv 匯出只會留下純檔名文字，超連結會遺失）。"""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid={gid}"
    resp = requests.get(url, timeout=20)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    return wb.active


def _fetch_market_table(market: str):
    """抓取並快取（24 小時）指定市場的 Game List 分頁，回傳 (normalize 過的表頭, 資料列)。"""
    now = time.time()
    cached = _market_table_cache.get(market)
    if cached and now - cached["fetched_at"] < _TABLE_CACHE_TTL_SECONDS:
        return cached["headers"], cached["rows"]
    cfg = _MARKET_DATA_SOURCES[market]
    ws = _fetch_xlsx(cfg["sheet_id"], cfg["gid"])
    all_rows = [_extract_row(r) for r in ws.iter_rows()]
    headers = [_normalize_header(text) for text, _ in all_rows[0]]
    data_rows = all_rows[1:]
    _market_table_cache[market] = {"headers": headers, "rows": data_rows, "fetched_at": now}
    return headers, data_rows


def _fetch_brazil_table():
    """Brazil 結構特殊：header 列不是第一列，前面還有幾列不分遊戲的 RGS/RNG 平台認證。
    回傳 (表頭, 逐遊戲資料列, 平台層級列)。"""
    now = time.time()
    cached = _brazil_table_cache.get("headers")
    if cached and now - _brazil_table_cache["fetched_at"] < _TABLE_CACHE_TTL_SECONDS:
        return _brazil_table_cache["headers"], _brazil_table_cache["rows"], _brazil_table_cache["platform_rows"]
    ws = _fetch_xlsx(_BRAZIL_SHEET_ID, _BRAZIL_GID)
    all_rows = [_extract_row(r) for r in ws.iter_rows()]
    header_idx = None
    for i, row in enumerate(all_rows):
        if "game id" in [_normalize_header(text) for text, _ in row]:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Brazil Game List 找不到表頭列（Game ID 欄位）")
    headers = [_normalize_header(text) for text, _ in all_rows[header_idx]]
    platform_rows = all_rows[:header_idx]
    data_rows = all_rows[header_idx + 1:]
    _brazil_table_cache.update(
        headers=headers, rows=data_rows, platform_rows=platform_rows, fetched_at=now
    )
    return headers, data_rows, platform_rows


def _resolve_brazil(doc_keyword: str, game_identifier: str = None):
    headers, rows, platform_rows = _fetch_brazil_table()
    lab_col = _find_col(headers, "lab")
    cert_col = _find_col(headers, "certification")
    if doc_keyword in ("rgs certificate", "rng certificate"):
        want = "rgs" if "rgs" in doc_keyword else "rng"
        for row in platform_rows:
            if lab_col is not None and lab_col < len(row) and row[lab_col][0].lower() == want:
                text, link = row[cert_col] if cert_col is not None and cert_col < len(row) else ("", None)
                if text:
                    parts = [f"{doc_keyword.upper()}: {text}"]
                    if link:
                        parts.append(link)
                    return True, "\n".join(parts)
        return False, None
    if not game_identifier:
        return False, None
    id_col = _find_col(headers, "game id")
    name_col = _find_col(headers, "game name")
    row = _find_row(rows, id_col, name_col, game_identifier)
    if row is None:
        return False, None
    text, link = row[cert_col] if cert_col is not None and cert_col < len(row) else ("", None)
    if not text:
        return False, None
    parts = [f"Certificate: {text}"]
    if link:
        parts.append(link)
    if lab_col is not None and lab_col < len(row) and row[lab_col][0]:
        parts.append(f"Lab: {row[lab_col][0]}")
    return True, "\n".join(parts)


def resolve_document(market: str, doc_keyword: str, game_identifier: str = None):
    """查表找出對應的認證文件資料。回傳 (found, text)：
    found=True 時 text 是要回覆的內容；found=False 時 text 一律是 None（查無資料，呼叫端自行決定轉人工文案）。"""
    try:
        if market == "Brazil":
            return _resolve_brazil(doc_keyword, game_identifier)
        if not game_identifier:
            return False, None
        headers, rows = _fetch_market_table(market)
        id_col = _find_col(headers, "gameid") or _find_col(headers, "game id")
        name_col = _find_col(headers, "name")
        cert_col = _find_col(headers, "certificate file")
        auth_col = _find_col(headers, "certification authority")
        hint = _KEYWORD_TO_COLUMN_HINT.get(doc_keyword, "certificate file")
        target_col = _find_col(headers, hint)
        row = _find_row(rows, id_col, name_col, game_identifier)
        if row is None or target_col is None:
            return False, None
        text, link = row[target_col] if target_col < len(row) else ("", None)
        if not text:
            return False, None
        parts = [f"{doc_keyword}: {text}"]
        if link:
            parts.append(link)
        if target_col == cert_col:
            # 問的本來就是通用認證檔：直接附上機構
            if auth_col is not None and auth_col < len(row) and row[auth_col][0]:
                parts.append(f"Certification Authority: {row[auth_col][0]}")
        elif cert_col is not None and cert_col < len(row):
            # 問的是特定文件類型（Help Files/DGOJ/HGC/Symbol Mapping/UKGC 等）：
            # 額外附上該市場通用的認證檔資訊做補充，比照文件範例（見「3.範例問答」GameID 659 那則）
            cert_text, cert_link = row[cert_col]
            if cert_text:
                auth_text = row[auth_col][0] if auth_col is not None and auth_col < len(row) else ""
                cert_line = f"Certificate file: {cert_text}" + (f" (Authority: {auth_text})" if auth_text else "")
                parts.append(cert_line)
                if cert_link:
                    parts.append(cert_link)
        return True, "\n".join(parts)
    except Exception as e:
        logger.error(f"[TadaCert] 查表失敗 market={market} keyword={doc_keyword}: {e}")
        return False, None


def find_inline_game_name(market: str, text: str):
    """在原始問句裡直接找有沒有出現該市場遊戲清單裡的完整遊戲名稱（不分大小寫、完整詞界比對，
    沿用 game_asset_service.find_games_by_exact_name 的作法）。extract_inline_game_id() 只認
    「GameID + 數字」這種明確格式，刻意不猜遊戲名稱，導致廠商在第一句話就講清楚市場+文件+遊戲
    名稱（例如「Fortune gems brazil certification」）時仍然會被追問一次遊戲。這裡額外用真正的
    遊戲清單比對一次，找到唯一一款相符的遊戲名稱就回傳，找不到、同時比對到多款、或該市場目前
    抓不到遊戲清單時回傳 None，維持原本的追問流程，不用猜的。"""
    from services.game_asset_service import find_games_by_exact_name
    try:
        if market == "Brazil":
            headers, rows, _ = _fetch_brazil_table()
            name_col = _find_col(headers, "game name")
        else:
            headers, rows = _fetch_market_table(market)
            name_col = _find_col(headers, "name")
    except Exception as e:
        logger.error(f"[TadaCert] 抓取 {market} 遊戲清單失敗（用於比對文字裡的遊戲名稱）：{e}")
        return None
    if name_col is None:
        return None
    name_list = [{"name": row[name_col][0]} for row in rows if name_col < len(row) and row[name_col][0]]
    matched = find_games_by_exact_name(text, name_list)
    if len(matched) == 1:
        return matched[0]
    return None


def is_whole_market_request(text: str) -> bool:
    lower = text.lower()
    return any(p.lower() in lower for p in _WHOLE_MARKET_PHRASES)


def market_reference_link(market: str, doc_keyword: str = None) -> str:
    """市場層級參考連結：該(市場,關鍵字)若屬於完全沒有逐遊戲資料的情況，優先給那個專屬資料夾連結；
    否則退回市場自己的備用資料夾連結，都沒有的話再退回整份 Game List 連結。"""
    if doc_keyword and (market, doc_keyword) in _NO_PER_GAME_DATA:
        return _NO_PER_GAME_DATA[(market, doc_keyword)]
    cfg = _MARKET_DATA_SOURCES.get(market)
    if cfg:
        return cfg["folder_link"] or cfg["gamelist_link"]
    return ""


def market_gamelist_link(market: str) -> str:
    """指定市場的 Game List 試算表連結（一定是試算表本身，不是認證文件的 Drive 資料夾）——
    給 TADA Gamelist 進階查詢用，例如廠商只問「Demo Link」但有講市場、沒講遊戲時，
    直接給該市場清單參考，因為 Game Demo 這類欄位只會存在於試算表裡，不在資料夾連結裡。"""
    cfg = _MARKET_DATA_SOURCES.get(market)
    if cfg:
        return cfg["gamelist_link"]
    if market in _SHARED_REPO_MARKETS:
        return _SHARED_REPO_MARKETS[market]
    if market == "Brazil":
        return f"https://docs.google.com/spreadsheets/d/{_BRAZIL_SHEET_ID}/edit"
    return ""


def whole_market_reply(market: str, is_chinese: bool, doc_keyword: str = None) -> str:
    link = market_reference_link(market, doc_keyword)
    if is_chinese:
        return f"請參考 {market} 認證文件總資料夾：\n{link}" if link else f"目前查詢不到 {market} 的相關資料，已為您轉接專人確認。"
    return f"Please refer to the {market} certification reference:\n{link}" if link else f"This certification data is currently unavailable for {market}, our team will assist you shortly."


def is_chinese_text(text: str) -> bool:
    """判斷訊息語言（含中文字元即視為中文），用來決定追問/回覆要用中文還是英文。"""
    return bool(_ZH_RE.search(text))


def detect_doc_query(text: str):
    """偵測訊息是否命中認證文件關鍵字。回傳 (doc_keyword, markets, doc_type) 或 None。
    doc_type 為 "no_data" 時 markets 固定回傳空清單（5d 情境不需要市場清單，直接轉人工）；
    只講泛稱詞（certificate/certification/認證/认证，見 _GENERIC_CERT_WORDS）、沒有更具體的
    詞組時，視同在問最常見的 certificate file，doc_type 回傳 "game"、markets 回傳
    _CERTIFICATE_FILE_MARKETS，呼叫端只需追問「哪個市場」即可。"""
    lower = text.lower()
    for kw in _NO_DATA_KEYWORDS:
        if kw in lower:
            return (kw, [], "no_data")
    for keywords, markets, doc_type in _DOC_KEYWORD_RULES:
        for kw in keywords:
            if kw in lower:
                return (kw, markets, doc_type)
    for kw in _GENERIC_CERT_WORDS:
        if kw in lower:
            return (kw, _CERTIFICATE_FILE_MARKETS, "game")
    return None


def _alias_in(alias: str, lower_text: str) -> bool:
    """英文別名用詞界比對，避免像 "uk" 是 "ukraine" 子字串這種誤配；中文沒有詞界概念，維持子字串比對。"""
    if alias.isascii():
        return bool(re.search(r'\b' + re.escape(alias) + r'\b', lower_text))
    return alias in lower_text


def detect_market_in_text(text: str) -> list:
    """從訊息裡找出有沒有明確提到市場名稱，回傳命中的市場清單。"""
    lower = text.lower()
    hits = []
    for market, aliases in _MARKET_ALIASES.items():
        if any(_alias_in(alias.lower(), lower) for alias in aliases):
            hits.append(market)
    return hits


def resolve_market_reply(text: str, candidates: list):
    """用使用者對追問的回覆文字，比對候選市場清單，回傳唯一命中的市場名稱；
    找不到或同時命中多個候選市場則回傳 None（維持追問狀態，讓呼叫端再問一次）。"""
    hits = [m for m in detect_market_in_text(text) if m in candidates]
    if len(hits) == 1:
        return hits[0]
    return None


def extract_game_identifier(text: str) -> str:
    """從追問回覆裡取出遊戲識別（GameID 或名稱）。廠商常直接打遊戲名稱而非純數字，
    所以不強制要求格式，直接回傳去除頭尾空白的原文。"""
    return text.strip()


# ── 訊息文案（依 is_chinese 決定中/英文版本）────────────────────────────────

def game_question(is_chinese: bool) -> str:
    return "請問是哪一款遊戲（GameID或遊戲名稱）？" if is_chinese else "Which game is this regarding? (GameID or game name)"


def vague_question(is_chinese: bool) -> str:
    """訊息只看得出跟認證有關，但完全對不到具體市場/文件類型時的最寬泛追問。"""
    return ("請問是哪個市場、需要哪一種認證文件呢？" if is_chinese
            else "Which market and which type of certification document do you need?")


def build_market_question(doc_keyword: str, candidates: list, is_chinese: bool) -> str:
    options = " / ".join(candidates)
    if is_chinese:
        return f"請問是哪個市場的 {doc_keyword}？（{options}）"
    return f"Which market's {doc_keyword} are you asking about? ({options})"


def market_mismatch_message(candidates: list, is_chinese: bool) -> str:
    options = " / ".join(candidates)
    if is_chinese:
        return f"沒有辨識到您指的市場，請從以下其中一個回覆：{options}"
    return f"We couldn't identify the market from your reply, please reply with one of: {options}"


def no_data_message(is_chinese: bool) -> str:
    return ("目前查詢不到相關認證資料，已為您轉接專人確認。" if is_chinese
            else "This certification data is currently unavailable, our team will assist you shortly.")


def no_per_game_data_message(is_chinese: bool, link: str) -> str:
    """該市場/文件類型目前完全沒有整併進逐遊戲資料（例如 Spain Help Files、Italy Game Project、
    Peru），誠實告知並退回專屬的整個市場資料夾連結，不要因為查不到就直接轉人工escalate。"""
    if is_chinese:
        return f"這份文件目前還沒有整併成逐遊戲資料，請參考以下連結：\n{link}"
    return f"This document isn't organized per game yet, please check the link below:\n{link}"


def vendor_escalate_message(is_chinese: bool) -> str:
    return ("這份文件依廠商各自產生，無法自動提供通用版本，已為您轉接專人確認。" if is_chinese
            else "This document is generated per vendor and has no generic version, our team will assist you shortly.")


def build_final_reply(market: str, doc_keyword: str, doc_type: str, is_chinese: bool, game_identifier: str = None) -> str:
    """market+doc_type（+game）都確定後：實際查表，找得到就回文件內容，找不到就誠實告知查無資料並轉人工。
    這裡會做網路請求（讀 Google Sheet XLSX），呼叫端須用 asyncio.to_thread 包起來，不要在事件迴圈裡直接呼叫。"""
    no_data_link = _NO_PER_GAME_DATA.get((market, doc_keyword))
    if no_data_link:
        return no_per_game_data_message(is_chinese, no_data_link)
    found, content = resolve_document(market, doc_keyword, game_identifier)
    if found:
        intro = "請見以下認證連結：" if is_chinese else "Please see the certification link below:"
        return f"{intro}\n{content}"
    return no_data_message(is_chinese)


def _timeout_reask_question(p) -> str:
    if p.missing == "market":
        candidates = json.loads(p.candidate_markets or "[]")
        options = " / ".join(candidates)
        if p.is_chinese:
            return f"請問您方才詢問的「{p.doc_keyword}」，是要查哪個市場的資料呢？（{options}）"
        return f"Regarding your earlier question about \"{p.doc_keyword}\", which market is this for? ({options})"
    if p.missing == "vague":
        return vague_question(p.is_chinese)
    if p.is_chinese:
        return f"請問您方才詢問的「{p.doc_keyword}」，是要查哪一款遊戲的資料呢？（GameID或遊戲名稱）"
    return f"Regarding your earlier question about \"{p.doc_keyword}\", which game is this for? (GameID or game name)"


# ── 追問待答狀態的資料庫存取 ──────────────────────────────────────────────

def get_pending(db, bot_id: int, chat_id: str, user_id: str):
    import models
    return db.query(models.TadaCertPending).filter(
        models.TadaCertPending.bot_id == bot_id,
        models.TadaCertPending.chat_id == chat_id,
        models.TadaCertPending.user_id == user_id,
    ).first()


def clear_pending(db, pending):
    db.delete(pending)
    db.commit()


def save_pending(db, bot_id: int, chat_id: str, user_id: str, prompt_message_id: int, missing: str,
                  doc_keyword: str, doc_type: str, is_chinese: bool, candidate_markets: list = None,
                  resolved_market: str = None, original_text: str = ""):
    """建立（或取代既有的）一筆待答狀態；每人同一時間只保留一筆，不做對話樹。"""
    import models
    existing = get_pending(db, bot_id, chat_id, user_id)
    if existing:
        db.delete(existing)
        db.flush()
    pending = models.TadaCertPending(
        bot_id=bot_id, chat_id=chat_id, user_id=user_id,
        prompt_message_id=prompt_message_id, missing=missing,
        doc_keyword=doc_keyword, doc_type=doc_type, is_chinese=is_chinese,
        candidate_markets=json.dumps(candidate_markets or []),
        resolved_market=resolved_market,
        original_text=original_text,
        expires_at=datetime.utcnow() + timedelta(minutes=PENDING_TTL_MINUTES),
    )
    db.add(pending)
    db.commit()
    return pending


# ── 逾時背景排程：每分鐘掃描一次，第一次逾時換句話重問，二次逾時轉人工並清除 ──────

async def cert_followup_timeout_loop():
    logger.info("[TadaCert] 追問逾時排程已啟動")
    while True:
        await asyncio.sleep(60)
        try:
            await _sweep_expired()
        except Exception as e:
            logger.error(f"[TadaCert] 逾時掃描例外: {e}")


async def _sweep_expired():
    import models
    from database import SessionLocal
    from services.telegram_service import bot_manager

    db = SessionLocal()
    try:
        now = datetime.utcnow()
        expired = db.query(models.TadaCertPending).filter(
            models.TadaCertPending.expires_at <= now
        ).all()
        for p in expired:
            try:
                await _handle_expired(db, bot_manager, p)
            except Exception as e:
                logger.error(f"[TadaCert] 處理 pending id={p.id} 逾時失敗: {e}")
    finally:
        db.close()


async def _handle_expired(db, bot_manager, p):
    from telegram import ForceReply

    if not p.reasked:
        question = _timeout_reask_question(p)
        sent = await asyncio.to_thread(
            bot_manager.send_message, p.bot_id, p.chat_id, question,
            reply_markup=ForceReply(selective=True),
        )
        p.reasked = True
        if sent is not None:
            p.prompt_message_id = sent.message_id
        p.expires_at = datetime.utcnow() + timedelta(minutes=PENDING_TTL_MINUTES)
        db.commit()
        logger.info(f"[TadaCert] pending id={p.id} 逾時，已重新追問一次")
    else:
        fallback = ("您好，人員將會協助確認，請稍後" if p.is_chinese
                    else "Hello, our team will assist you shortly. Please wait.")
        await asyncio.to_thread(bot_manager.send_message, p.bot_id, p.chat_id, fallback)
        db.delete(p)
        db.commit()
        logger.info(f"[TadaCert] pending id={p.id} 二次逾時，已轉人工並清除待答狀態")
