"""
知識庫 AI 服務
- 文件段落儲存於 SQLite（KnowledgeChunk / TeamsKnowledgeChunk）
- 以關鍵字評分找出相關段落，再呼叫 Claude 產生答案
- 完全不使用 ChromaDB/ONNX，記憶體佔用極低
"""
import os
import asyncio
import logging
from datetime import date
from typing import Optional

logger = logging.getLogger(__name__)

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")

_anthropic_client = None


def get_anthropic_client():
    global _anthropic_client
    if _anthropic_client is None:
        import anthropic
        _anthropic_client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))
    return _anthropic_client


# ── 文件解析 ────────────────────────────────────────────────────────────────

def _extract_text(file_path: str, ext: str) -> str:
    if ext == ".pdf":
        import PyPDF2
        text = []
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text.append(page.extract_text() or "")
        return "\n".join(text)

    elif ext in (".doc", ".docx"):
        from docx import Document
        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs)

    elif ext in (".xls", ".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                lines.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(lines)

    elif ext in (".txt", ".csv", ".md"):
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    return ""


def _parse_excel_qa(file_path: str) -> list[dict]:
    """讀取 Excel，A 欄=問題、B 欄=答案，自動轉為 Q&A 清單。"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    qas = []
    skip_headers = {"問題", "問題內容", "Q", "question", "題目"}
    skip_answers = {"回覆", "回覆內容", "A", "answer", "答案"}
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if len(row) < 2:
                continue
            q = str(row[0]).strip() if row[0] is not None else ""
            a = str(row[1]).strip() if row[1] is not None else ""
            if not q or not a:
                continue
            if q in skip_headers and a in skip_answers:
                continue
            qas.append({"question": q, "keywords": "", "answer": a})
    return qas


def _is_cjk(text: str) -> bool:
    cjk = sum(1 for c in text if '一' <= c <= '鿿')
    return len(text) > 0 and cjk / len(text) > 0.2


def parse_qa_text(text: str) -> list[dict]:
    """
    解析 Q&A 格式文字，回傳 [{"question": ..., "keywords": ..., "answer": ...}, ...]
    格式：
    Q1: 問題標題
    問題的其他說法或關鍵字（可選，接在 Q 行之後、A 行之前）
    A1: 回答內容
    ------------------------------------------------------------
    Q2: ...
    """
    import re
    blocks = re.split(r'\n-{10,}\n', text)
    qas = []
    qa_re = re.compile(
        r'Q\d*[:：]\s*(.+?)\n(.*?)A\d*[:：]\s*(.+)',
        re.DOTALL | re.IGNORECASE
    )
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        m = qa_re.match(block)
        if m:
            question = m.group(1).strip()
            keywords = m.group(2).strip()
            answer = m.group(3).strip()
            qas.append({"question": question, "keywords": keywords, "answer": answer})
    return qas


def _chunk_text(text: str, chunk_size: int = 400, overlap: int = 40):
    import re

    # 偵測 Q&A 格式（以 --- 分隔線分段），每組 Q&A 保留完整
    if re.search(r'\n-{10,}\n', text):
        blocks = re.split(r'\n-{10,}\n', text)
        chunks = [b.strip() for b in blocks if b.strip()]
        logger.info(f"偵測到 Q&A 格式，共 {len(chunks)} 個問答段落")
        return chunks

    if _is_cjk(text):
        char_size, char_overlap = 400, 50
        chunks, i = [], 0
        while i < len(text):
            chunks.append(text[i:i + char_size])
            i += char_size - char_overlap
    else:
        words = text.split()
        chunks, i = [], 0
        while i < len(words):
            chunks.append(" ".join(words[i:i + chunk_size]))
            i += chunk_size - overlap
    return [c for c in chunks if c.strip()]


# ── 文件處理（上傳時呼叫）──────────────────────────────────────────────────

async def process_document(file_path: str, ext: str, collection_name: str,
                           doc_id: int = None, bot_id: int = None,
                           db=None, platform: str = "telegram"):
    """
    解析文件並將段落存入 SQLite。
    collection_name 保留供相容，實際不使用 ChromaDB。
    """
    text = await asyncio.to_thread(_extract_text, file_path, ext)
    if not text.strip():
        raise ValueError("無法從文件中提取文字")

    chunks = _chunk_text(text)

    if db is not None and doc_id is not None and bot_id is not None:
        import models
        if platform == "teams":
            ChunkModel = models.TeamsKnowledgeChunk
        elif platform == "copilot":
            ChunkModel = models.CopilotKnowledgeChunk
        else:
            ChunkModel = models.KnowledgeChunk
        for i, chunk in enumerate(chunks):
            db.add(ChunkModel(doc_id=doc_id, bot_id=bot_id,
                              chunk_text=chunk, chunk_index=i))

        # Telegram 平台：同步寫入 KnowledgeQA
        if platform == "telegram":
            # Excel 直接以 A欄=Q、B欄=A 解析，其餘走 Q&A 文字格式
            if ext in (".xls", ".xlsx"):
                qas = _parse_excel_qa(file_path)
            else:
                qas = parse_qa_text(text)
            for i, qa in enumerate(qas):
                db.add(models.KnowledgeQA(
                    doc_id=doc_id, bot_id=bot_id,
                    question=qa["question"],
                    keywords=qa["keywords"],
                    answer=qa["answer"],
                    order_index=i,
                ))
            if qas:
                logger.info(f"文件 doc_id={doc_id} 解析到 {len(qas)} 組 Q&A")

        db.commit()
        logger.info(f"文件 doc_id={doc_id} 共 {len(chunks)} 個段落已存入 SQLite")


def delete_document_vectors(collection_name: str):
    """舊介面保留，ChromaDB 已移除，此處為空操作"""
    pass


# ── 關鍵字評分搜尋 ──────────────────────────────────────────────────────────

def _extract_tokens(text: str) -> set[str]:
    """
    從問題中提取搜尋 token：
    - 英文：空白分詞；若為長度 >4 的連寫詞（如 APIURL）再拆 3-4 字子串
    - 中文：連續漢字序列拆成 2~4 字片段（小數點、支援、位數…）
    """
    import re
    tokens: set[str] = set()

    # 英文空白分詞
    for w in text.lower().split():
        w = re.sub(r'[^\w]', '', w)
        if len(w) <= 1:
            continue
        tokens.add(w)
        # 長連寫詞（如 apiurl）也拆成 3-4 字子串，讓 api、url 各別匹配
        if len(w) > 4:
            for n in (3, 4):
                for i in range(len(w) - n + 1):
                    tokens.add(w[i:i + n])

    # 提取連續漢字序列，拆成 2~4 字片段
    cjk_seqs = re.findall(r'[一-鿿]+', text)
    for seq in cjk_seqs:
        for n in range(2, min(5, len(seq) + 1)):
            for i in range(len(seq) - n + 1):
                tokens.add(seq[i:i + n])

    return tokens


def _score_chunks(question: str, chunks) -> list[str]:
    """所有 chunk 直接送給 Claude，由 Claude 判斷語意相關性"""
    return [c.chunk_text for c in chunks]


# ── 知識庫查詢（Telegram）────────────────────────────────────────────────────

def query_knowledge(bot_id: int, question: str, db=None) -> Optional[tuple]:
    """
    db 參數保留供除錯端點直接呼叫；
    從 asyncio.to_thread 呼叫時請不要傳 db，函式會自行建立 session。
    """
    import models
    from database import SessionLocal

    own_db = db is None
    if own_db:
        db = SessionLocal()

    try:
        logger.info(f"Telegram Bot {bot_id} 開始知識庫查詢：「{question[:50]}」")

        docs = db.query(models.KnowledgeDoc).filter(
            models.KnowledgeDoc.bot_id == bot_id,
            models.KnowledgeDoc.is_enabled == True
        ).all()

        if not docs:
            logger.info(f"Bot {bot_id} 無啟用的知識庫文件")
            return None

        doc_ids = [d.id for d in docs]
        all_chunks = db.query(models.KnowledgeChunk).filter(
            models.KnowledgeChunk.doc_id.in_(doc_ids)
        ).all()

        if not all_chunks:
            logger.warning(f"Bot {bot_id} 知識庫文件無段落資料（可能需重新上傳）")
            return None

        top_chunks = _score_chunks(question, all_chunks)
        return _call_claude(question, top_chunks, bot_id)
    finally:
        if own_db:
            db.close()


# ── 知識庫查詢（Teams）───────────────────────────────────────────────────────

def query_teams_knowledge(bot_id: int, question: str, db=None) -> Optional[tuple]:
    import models
    from database import SessionLocal

    own_db = db is None
    if own_db:
        db = SessionLocal()

    try:
        logger.info(f"Teams Bot {bot_id} 開始知識庫查詢：「{question[:50]}」")

        docs = db.query(models.TeamsKnowledgeDoc).filter(
            models.TeamsKnowledgeDoc.bot_id == bot_id,
            models.TeamsKnowledgeDoc.is_enabled == True
        ).all()

        if not docs:
            return None

        doc_ids = [d.id for d in docs]
        all_chunks = db.query(models.TeamsKnowledgeChunk).filter(
            models.TeamsKnowledgeChunk.doc_id.in_(doc_ids)
        ).all()

        if not all_chunks:
            return None

        top_chunks = _score_chunks(question, all_chunks)
        return _call_claude(question, top_chunks, bot_id)
    finally:
        if own_db:
            db.close()


# ── 呼叫 Claude ──────────────────────────────────────────────────────────────

def _call_claude(question: str, chunks: list[str], bot_id: int) -> Optional[tuple]:
    context = "\n\n".join(chunks)
    logger.info(f"Bot {bot_id} 呼叫 Claude，context {len(context)} 字元")
    try:
        client = get_anthropic_client()
        system_prompt = (
            "你是一個問答機器人。請根據以下知識庫內容，直接回答問題的答案，"
            "不要加入任何額外補充、說明或開場白。"
            "重要規則：\n"
            "1. 只有在知識庫中有【直接對應的答案】時才回答，不可自行推斷、整合或延伸。\n"
            "2. 若訊息是申請單、表單或提交資料（特徵：多行編號欄位如「1. 商戶名字：xxx」、「2. 幣種：MYR」等結構），"
            "視為提交資料而非提問，請回覆找不到相關資訊。\n"
            "3. 請使用與問題完全相同的語言回答"
            "（問題為繁體中文→繁體中文、英文→英文、簡體中文→簡體中文、其他語言同理）。\n"
            "4. 若知識庫中沒有直接對應的資訊，僅以問題的語言簡短告知找不到相關資訊即可。"
        )
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            system=[
                {
                    "type": "text",
                    "text": system_prompt,
                },
                {
                    "type": "text",
                    "text": f"知識庫內容：\n{context}",
                    "cache_control": {"type": "ephemeral"},
                },
            ],
            messages=[{
                "role": "user",
                "content": f"問題：{question}\n\n答案："
            }],
            extra_headers={"anthropic-beta": "prompt-caching-2024-07-31"},
        )
        reply = message.content[0].text.strip()
        cache_read = getattr(message.usage, "cache_read_input_tokens", 0) or 0
        cache_write = getattr(message.usage, "cache_creation_input_tokens", 0) or 0
        logger.info(
            f"Bot {bot_id} Claude 回覆成功，tokens: in={message.usage.input_tokens} "
            f"cache_read={cache_read} cache_write={cache_write}"
        )
        in_tok, out_tok = message.usage.input_tokens, message.usage.output_tokens
        # 若 Claude 判斷找不到答案，回傳 (None, tokens) 讓上層仍能記錄用量
        NO_ANSWER_SIGNALS = [
            "找不到相關資訊", "無法回答", "沒有相關",
            "I cannot find", "no relevant", "not found", "no information",
            "找不到相关", "无法回答",  # 簡體中文
        ]
        if any(s in reply for s in NO_ANSWER_SIGNALS):
            logger.info(f"Bot {bot_id} Claude 表示找不到答案，觸發 fallback")
            return None, in_tok, out_tok, cache_read, cache_write
        return reply, in_tok, out_tok, cache_read, cache_write
    except Exception as e:
        logger.error(f"Bot {bot_id} Claude API 失敗：{e}", exc_info=True)
        return None, 0, 0, 0, 0


# ── 使用量記錄 ────────────────────────────────────────────────────────────────

def record_usage(bot_id: int, input_tokens: int, output_tokens: int, db,
                 cache_read_tokens: int = 0, cache_write_tokens: int = 0):
    import models
    today = date.today().isoformat()
    # 實際計費等效 token：input + output + cache_read×10% + cache_write×125%
    effective = input_tokens + output_tokens + round(cache_read_tokens * 0.1) + round(cache_write_tokens * 1.25)
    stat = db.query(models.UsageStat).filter(
        models.UsageStat.bot_id == bot_id,
        models.UsageStat.date == today
    ).first()
    if stat:
        stat.input_tokens += input_tokens
        stat.output_tokens += output_tokens
        stat.cache_read_tokens += cache_read_tokens
        stat.cache_write_tokens += cache_write_tokens
        stat.total_tokens += effective
        stat.request_count += 1
    else:
        stat = models.UsageStat(
            bot_id=bot_id, date=today,
            input_tokens=input_tokens, output_tokens=output_tokens,
            cache_read_tokens=cache_read_tokens, cache_write_tokens=cache_write_tokens,
            total_tokens=effective, request_count=1,
        )
        db.add(stat)
    db.commit()
